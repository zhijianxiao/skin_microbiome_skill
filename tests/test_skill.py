"""Tests for skinmicrobiome v0.4.0 — Boolean keyword + scoring."""

import os
import json
import tempfile
from pathlib import Path
from unittest.mock import patch, MagicMock

import pytest

from skill import (
    SkinMicrobiomeSkill,
    SpeciesManager,
    SearchEngine,
    ENAClient,
    Exporter,
    detect_platform,
)

# ═══════════════════════════════════════════════════════════════════════
# Fake NCBI data (matching bioSkills patterns)
# ═══════════════════════════════════════════════════════════════════════

_FAKE_PUBMED = {
    "PubmedArticle": [
        {
            "MedlineCitation": {
                "PMID": "40000001",
                "Article": {
                    "ArticleTitle": "Characterization of the equine skin microbiome in healthy horses",
                    "AuthorList": [
                        {"LastName": "Smith", "ForeName": "John"},
                        {"LastName": "Brown", "ForeName": "Alice"},
                    ],
                    "Journal": {
                        "Title": "Veterinary Dermatology",
                        "JournalIssue": {"PubDate": {"Year": "2025"}},
                    },
                    "ELocationID": [
                        {"EIdType": "doi", "value": "10.1111/vde.13001"},
                    ],
                    "Abstract": {
                        "AbstractText": [
                            "We collected skin swabs from the ear, axilla, and groin",
                            "of 20 healthy horses. Metagenomic sequencing revealed",
                            "diverse bacterial communities.",
                        ]
                    },
                },
            }
        },
        {
            "MedlineCitation": {
                "PMID": "40000002",
                "Article": {
                    "ArticleTitle": "Skin microbiota of laboratory mice — dorsal and ventral sites",
                    "AuthorList": [{"LastName": "Lee", "ForeName": "Sarah"}],
                    "Journal": {
                        "Title": "Lab Animal",
                        "JournalIssue": {"PubDate": {"Year": "2024"}},
                    },
                    "ELocationID": [
                        {"EIdType": "doi", "value": "10.1038/laban.2024.001"},
                    ],
                    "Abstract": {
                        "AbstractText": [
                            "We characterized the dorsal and ventral skin microbiome",
                            "of C57BL/6 mice using 16S rRNA sequencing.",
                        ]
                    },
                },
            }
        },
        {  # HUMAN — should be excluded
            "MedlineCitation": {
                "PMID": "40000003",
                "Article": {
                    "ArticleTitle": "Human skin microbiome in patients with atopic dermatitis",
                    "AuthorList": [{"LastName": "Kim", "ForeName": "Daniel"}],
                    "Journal": {
                        "Title": "J Invest Dermatol",
                        "JournalIssue": {"PubDate": {"Year": "2023"}},
                    },
                    "ELocationID": [
                        {"EIdType": "doi", "value": "10.1016/jid.2023.01.001"},
                    ],
                    "Abstract": {
                        "AbstractText": [
                            "We enrolled 50 human patients with atopic dermatitis...",
                        ]
                    },
                },
            }
        },
        {  # GUT, not skin — should be excluded
            "MedlineCitation": {
                "PMID": "40000004",
                "Article": {
                    "ArticleTitle": "Gut microbiota of horses fed different diets",
                    "AuthorList": [{"LastName": "Jones", "ForeName": "Mike"}],
                    "Journal": {
                        "Title": "Equine Vet J",
                        "JournalIssue": {"PubDate": {"Year": "2022"}},
                    },
                    "Abstract": {
                        "AbstractText": ["We analyzed fecal samples from 30 horses..."],
                    },
                },
            }
        },
        {
            "MedlineCitation": {
                "PMID": "40000005",
                "Article": {
                    "ArticleTitle": "The cutaneous microbiome of dogs with allergic dermatitis",
                    "AuthorList": [
                        {"LastName": "Park", "ForeName": "Eve"},
                        {"LastName": "Wilson", "ForeName": "Tom"},
                    ],
                    "Journal": {
                        "Title": "Vet Dermatol",
                        "JournalIssue": {"PubDate": {"Year": "2025"}},
                    },
                    "Abstract": {
                        "AbstractText": [
                            "Skin swabs from the paw and interdigital spaces of",
                            "atopic dogs were sequenced.",
                        ]
                    },
                },
            }
        },
    ]
}

# ELink PubMed → BioProject mapping
_FAKE_ELINK_PUBMED_BP = [
    {"IdList": ["40000001"], "LinkSetDb": [
        {"DbTo": "bioproject", "LinkName": "pubmed_bioproject",
         "Link": [{"Id": "PRJNA123456"}]}
    ]},
    {"IdList": ["40000002"], "LinkSetDb": [
        {"DbTo": "bioproject", "LinkName": "pubmed_bioproject",
         "Link": []}
    ]},
    {"IdList": ["40000003"], "LinkSetDb": []},
    {"IdList": ["40000004"], "LinkSetDb": []},
    {"IdList": ["40000005"], "LinkSetDb": [
        {"DbTo": "bioproject", "LinkName": "pubmed_bioproject",
         "Link": [{"Id": "PRJEB998877"}]}
    ]},
]

# ELink PubMed → SRA mapping
_FAKE_ELINK_PUBMED_SRA = [
    {"IdList": ["40000001"], "LinkSetDb": [
        {"DbTo": "sra", "LinkName": "pubmed_sra",
         "Link": [{"Id": "1234567"}, {"Id": "1234568"}]}
    ]},
    {"IdList": ["40000002"], "LinkSetDb": []},
    {"IdList": ["40000003"], "LinkSetDb": []},
    {"IdList": ["40000004"], "LinkSetDb": []},
    {"IdList": ["40000005"], "LinkSetDb": [
        {"DbTo": "sra", "LinkName": "pubmed_sra",
         "Link": [{"Id": "9876543"}]}
    ]},
]


def _entrez_read_mock(handle):
    name = str(getattr(handle, "name", ""))
    if "esearch" in name:
        return {"IdList": ["40000001", "40000002", "40000003", "40000004", "40000005"],
                "Count": "5"}
    if "elink" in name.lower():
        ids = getattr(handle, "_fake_ids", None)
        if ids and "40000001" in ids:
            return _FAKE_ELINK_PUBMED_BP
        if ids and "sra" in str(getattr(handle, "_fake_db", "")):
            return _FAKE_ELINK_PUBMED_SRA
        return _FAKE_ELINK_PUBMED_BP
    return _FAKE_PUBMED


@pytest.fixture
def entrez_mock():
    """Mock all Bio.Entrez calls with bioSkills-pattern data."""
    with patch("Bio.Entrez.esearch") as m_es, \
         patch("Bio.Entrez.efetch") as m_ef, \
         patch("Bio.Entrez.elink") as m_el, \
         patch("Bio.Entrez.read", side_effect=_entrez_read_mock), \
         patch("Bio.Entrez.esummary") as m_esu:
        m_es.return_value = MagicMock(name="esearch")
        m_es.return_value.read.return_value = ""
        m_ef.return_value = MagicMock(name="efetch")
        m_ef.return_value.read.return_value = ""
        m_el.return_value = MagicMock(name="elink")
        m_el.return_value._fake_ids = ["40000001","40000002","40000003","40000004","40000005"]
        m_esu.return_value = MagicMock(name="esummary")
        m_esu.return_value.read.return_value = ""
        yield


@pytest.fixture
def species_terms():
    return {
        "canonical": "Equus caballus",
        "common": "Horse",
        "synonyms": [],
        "genera": [],
        "parent_terms": ["Perissodactyla", "Mammalia"],
        "all": ["horse", "horses", "Equus caballus", "Perissodactyla", "Mammalia"],
        "boolean_clause": '("horse" OR "horses" OR "Equus caballus" OR "Perissodactyla" OR "Mammalia")',
    }


@pytest.fixture
def search_engine():
    return SearchEngine({
        "search": {"max_results": 100, "timeout": 30, "email": "test@test.com"},
    })


@pytest.fixture
def species_manager():
    return SpeciesManager()


@pytest.fixture
def exporter():
    return Exporter({"export": {"default_format": "xlsx", "output_dir": "./output"}})


@pytest.fixture
def sample_articles():
    return [
        {
            "pmid": "10001",
            "title": "Equine skin microbiome diversity in healthy horses",
            "authors": "Alice Chen; Bob Li",
            "year": "2024",
            "doi": "10.1000/equine.2024.001",
            "abstract": "Skin swabs from horse ear and back revealed diverse microbiota.",
            "species": "Equus caballus",
            "bioproject_ids": ["PRJNA123456"],
            "sampling_site": "Ear",
        },
        {
            "pmid": "10002",
            "title": "Mouse dorsal skin microbiota changes with age",
            "authors": "Charlie Wang",
            "year": "2023",
            "doi": "",
            "abstract": "Dorsal skin of C57BL/6 mice was sampled.",
            "species": "Mus musculus",
            "bioproject_ids": [],
            "sampling_site": "Back",
        },
        {
            "pmid": "10003",
            "title": "Canine cutaneous fungal communities in atopic dogs",
            "authors": "Diana Park; Eve Kim",
            "year": "2025",
            "doi": "10.1000/canine.2025.003",
            "abstract": "Paw and interdigital skin microbiome of dogs with atopic dermatitis.",
            "species": "Canis lupus familiaris",
            "bioproject_ids": ["PRJEB998877"],
            "sampling_site": "Paw",
        },
    ]


# ═══════════════════════════════════════════════════════════════════════
# 1. Species: local DB + NCBI Taxonomy fallback
# ═══════════════════════════════════════════════════════════════════════

class TestSpeciesNormalization:

    def test_ma_to_equus_caballus(self, species_manager):
        result = species_manager.normalize("马")
        assert result["latin"] == "Equus caballus"

    def test_shu_to_mus_musculus(self, species_manager):
        result = species_manager.normalize("鼠")
        assert result["latin"] == "Mus musculus"

    def test_xiaoshu_to_mus_musculus(self, species_manager):
        result = species_manager.normalize("小鼠")
        assert result["latin"] == "Mus musculus"

    def test_english_mouse(self, species_manager):
        result = species_manager.normalize("mouse")
        assert result["latin"] == "Mus musculus"

    def test_latin_direct(self, species_manager):
        result = species_manager.normalize("Equus caballus")
        assert result["chinese"] == "马"

    def test_case_insensitive(self, species_manager):
        result = species_manager.normalize("MUS MUSCULUS")
        assert result["latin"] == "Mus musculus"

    def test_unknown_returns_none(self, species_manager):
        assert species_manager.normalize("外星生物XYZ") is None

    def test_list_animals(self, species_manager):
        animals = species_manager.list_animals()
        latins = [a["latin"] for a in animals]
        assert "Equus caballus" in latins
        assert "Mus musculus" in latins

    def test_get_info(self, species_manager):
        info = species_manager.get_info("狗")
        assert info["latin"] == "Canis lupus familiaris"

    def test_get_info_unknown(self, species_manager):
        info = species_manager.get_info("火星细菌")
        assert "error" in info

    def test_search_by_keyword(self, species_manager):
        results = species_manager.search("马")
        assert any(r["latin"] == "Equus caballus" for r in results)

    def test_ncbi_taxonomy_fallback(self, species_manager):
        """Dynamic NCBI Taxonomy lookup for species not in local DB."""
        from Bio import Entrez
        fake_esearch = {"IdList": ["9999"], "Count": "1"}
        fake_esummary = [{"ScientificName": "Testus organismus",
                           "CommonName": "test bug"}]
        with patch("Bio.Entrez.esearch") as m_es, \
             patch("Bio.Entrez.read") as m_read, \
             patch("Bio.Entrez.esummary") as m_esu:
            m_es.return_value = MagicMock()
            m_read.side_effect = [fake_esearch, fake_esummary]
            m_esu.return_value = MagicMock()
            result = species_manager._ncbi_taxonomy_lookup("test bug")
            assert result is not None
            assert result["latin"] == "Testus organismus"
            assert result["taxon_id"] == "9999"


# ═══════════════════════════════════════════════════════════════════════
# 2. Search: ELink BioProject
# ═══════════════════════════════════════════════════════════════════════

class TestSearchEngine:

    def test_search_returns_articles(self, entrez_mock, search_engine):
        articles = search_engine.search(species_terms=species_terms)
        assert isinstance(articles, list)
        assert len(articles) >= 1
        for art in articles:
            assert "title" in art
            assert "sampling_site" in art
            assert "bioproject_ids" in art

    def test_human_patients_excluded(self, entrez_mock, search_engine):
        articles = search_engine.search(species_terms=species_terms)
        pmids = [a["pmid"] for a in articles]
        assert "40000003" not in pmids, "Human patient article must be excluded"

    def test_non_skin_excluded(self, entrez_mock, search_engine):
        articles = search_engine.search(species_terms=species_terms)
        pmids = [a["pmid"] for a in articles]
        assert "40000004" not in pmids, "Gut microbiota article must be excluded"

    def test_skin_articles_present(self, entrez_mock, search_engine):
        articles = search_engine.search(species_terms=species_terms)
        pmids = [a["pmid"] for a in articles]
        assert "40000001" in pmids
        assert "40000005" in pmids

    def test_elink_bioproject_assignment(self, entrez_mock, search_engine):
        """ELink should assign PRJNA123456 to article 40000001."""
        articles = search_engine.search(species_terms=species_terms)
        art1 = next((a for a in articles if a["pmid"] == "40000001"), None)
        assert art1 is not None
        assert "PRJNA123456" in art1["bioproject_ids"]

    def test_elink_empty_bioproject(self, entrez_mock, search_engine):
        """Article 40000002 has no BioProject via ELink."""
        articles = search_engine.search(species_terms=species_terms)
        art2 = next((a for a in articles if a["pmid"] == "40000002"), None)
        assert art2 is not None
        assert art2["bioproject_ids"] == []

    def test_sampling_site_extraction(self, entrez_mock, search_engine):
        articles = search_engine.search(species_terms=species_terms)
        art1 = next((a for a in articles if a["pmid"] == "40000001"), None)
        assert art1 is not None
        assert art1["sampling_site"] != "N/A"
        assert art1["sampling_site"] in ("Ear", "Axilla", "Groin")

    def test_sampling_site_n_a(self, entrez_mock, search_engine):
        articles = search_engine.search(species_terms=species_terms)
        art2 = next((a for a in articles if a["pmid"] == "40000002"), None)
        assert art2 is not None
        assert art2["sampling_site"] != "N/A"

    def test_search_with_extra_keywords(self, entrez_mock, search_engine):
        articles = search_engine.search(
            species_terms=species_terms,
            extra_keywords="dermatitis",
        )
        assert isinstance(articles, list)
        assert len(articles) >= 1

    def test_extract_bioproject_regex_fallback(self, search_engine):
        ids = search_engine.extract_bioproject_ids({
            "title": "Study PRJNA998877",
            "abstract": "Also PRJEB112233 in the data.",
        })
        assert "PRJNA998877" in ids
        assert "PRJEB112233" in ids

    def test_extract_ena_ids_all(self, search_engine):
        ids = search_engine.extract_ena_ids({
            "title": "SRR11111 ERR22222 SAMN33333",
            "abstract": "",
        })
        assert "SRR11111" in ids
        assert "ERR22222" in ids
        assert "SAMN33333" in ids

    def test_query_builder_fields(self, search_engine):
        q = search_engine._build_query("Equus caballus", "")
        assert "skin microbiome" in q
        assert "Equus caballus" in q
        assert "human[mesh]" in q
        assert "[title/abstract]" in q


# ═══════════════════════════════════════════════════════════════════════
# 3. Export: 7 columns, N/A for missing
# ═══════════════════════════════════════════════════════════════════════

_EXPECTED_HEADERS = ["文献标题", "作者", "年份", "DOI", "实验物种", "宏基因组数据集", "取样部位"]


class TestExporter:

    def test_exactly_7_columns(self, exporter, sample_articles):
        with tempfile.TemporaryDirectory() as tmpdir:
            fp = os.path.join(tmpdir, "test.xlsx")
            exporter.to_excel(sample_articles, fp)
            from openpyxl import load_workbook
            wb = load_workbook(fp)
            ws = wb.active
            headers = [ws.cell(row=1, column=c).value for c in range(1, 9)]
            assert headers == _EXPECTED_HEADERS
            assert ws.cell(row=1, column=8).value is None
            wb.close()

    def test_missing_doi_is_n_a(self, exporter, sample_articles):
        with tempfile.TemporaryDirectory() as tmpdir:
            fp = os.path.join(tmpdir, "test.xlsx")
            exporter.to_excel(sample_articles, fp)
            from openpyxl import load_workbook
            wb = load_workbook(fp)
            ws = wb.active
            assert ws.cell(row=3, column=4).value == "N/A"  # article 2 has no DOI
            wb.close()

    def test_missing_bioproject_is_n_a(self, exporter, sample_articles):
        with tempfile.TemporaryDirectory() as tmpdir:
            fp = os.path.join(tmpdir, "test.xlsx")
            exporter.to_excel(sample_articles, fp)
            from openpyxl import load_workbook
            wb = load_workbook(fp)
            ws = wb.active
            assert ws.cell(row=3, column=6).value == "N/A"
            wb.close()

    def test_all_rows_complete(self, exporter, sample_articles):
        with tempfile.TemporaryDirectory() as tmpdir:
            fp = os.path.join(tmpdir, "test.xlsx")
            exporter.to_excel(sample_articles, fp)
            from openpyxl import load_workbook
            wb = load_workbook(fp)
            ws = wb.active
            assert ws.max_row == 4  # header + 3 data
            for row in range(2, 5):
                for col in range(1, 8):
                    assert ws.cell(row=row, column=col).value is not None
            wb.close()

    def test_csv_export(self, exporter, sample_articles):
        with tempfile.TemporaryDirectory() as tmpdir:
            fp = os.path.join(tmpdir, "test.csv")
            exporter.to_csv(sample_articles, fp)
            content = Path(fp).read_text(encoding="utf-8-sig")
            assert "N/A" in content
            for h in _EXPECTED_HEADERS:
                assert h in content

    def test_json_export(self, exporter, sample_articles):
        with tempfile.TemporaryDirectory() as tmpdir:
            fp = os.path.join(tmpdir, "test.json")
            exporter.to_json(sample_articles, fp)
            with open(fp, "r", encoding="utf-8") as f:
                data = json.load(f)
            assert len(data) == 3

    def test_empty_list_export(self, exporter):
        with tempfile.TemporaryDirectory() as tmpdir:
            fp = os.path.join(tmpdir, "empty.xlsx")
            exporter.to_excel([], fp)
            assert os.path.getsize(fp) > 0

    def test_excel_has_autofilter(self, exporter, sample_articles):
        with tempfile.TemporaryDirectory() as tmpdir:
            fp = os.path.join(tmpdir, "filtered.xlsx")
            exporter.to_excel(sample_articles, fp)
            from openpyxl import load_workbook
            wb = load_workbook(fp)
            ws = wb.active
            assert ws.auto_filter.ref is not None
            wb.close()


# ═══════════════════════════════════════════════════════════════════════
# 4. ENA: Entrez ESummary BioProject + ELink SRA
# ═══════════════════════════════════════════════════════════════════════

class TestENAClient:

    def test_resolve_bioproject_via_esummary(self):
        client = ENAClient()
        fake_summary = [{"title": "Equine skin metagenome", "Project_Title": "Horse skin"}]
        with patch("Bio.Entrez.esummary") as m_esu, \
             patch("Bio.Entrez.read", return_value=fake_summary):
            m_esu.return_value = MagicMock()
            result = client.resolve_bioproject("PRJNA123456")
            assert result == "Equine skin metagenome"

    def test_resolve_bioproject_n_a_on_error(self):
        client = ENAClient()
        with patch("Bio.Entrez.esummary", side_effect=Exception("fail")), \
             patch("skill.ena._ena_get", return_value=None):
            result = client.resolve_bioproject("PRJNA000000")
            assert result == "N/A"

    def test_get_sra_runs_for_bioproject(self):
        client = ENAClient()
        fake_elink = [{"IdList": ["PRJNA123"], "LinkSetDb": [
            {"DbTo": "sra", "Link": [{"Id": "9999"}]}
        ]}]
        with patch("Bio.Entrez.elink") as m_el, \
             patch("Bio.Entrez.read", return_value=fake_elink), \
             patch("Bio.Entrez.efetch") as m_ef:
            m_el.return_value = MagicMock()
            m_ef.return_value = MagicMock()
            m_ef.return_value.read.return_value = b"Run,spots,bases\nSRR11111,1000,150000\nSRR22222,2000,300000\n"
            runs = client.get_sra_runs_for_bioproject("PRJNA123")
            assert "SRR11111" in runs


# ═══════════════════════════════════════════════════════════════════════
# 5. Integration
# ═══════════════════════════════════════════════════════════════════════

class TestIntegration:

    def test_init_loads_all_modules(self):
        skill = SkinMicrobiomeSkill(config_path="config.yaml")
        assert skill.species_manager is not None
        assert skill.search_engine is not None
        assert skill.ena_client is not None
        assert skill.exporter is not None

    def test_get_species_info_ma(self):
        skill = SkinMicrobiomeSkill(config_path="config.yaml")
        assert skill.get_species_info("马")["latin"] == "Equus caballus"

    def test_get_species_info_shu(self):
        skill = SkinMicrobiomeSkill(config_path="config.yaml")
        assert skill.get_species_info("鼠")["latin"] == "Mus musculus"

    def test_detect_platform(self):
        plat = detect_platform()
        assert plat in ("linux", "windows", "macos")

    def test_search_delegates(self, entrez_mock):
        skill = SkinMicrobiomeSkill(config_path="config.yaml")
        results = skill.search(species_terms=species_terms)
        assert isinstance(results, list)
        assert len(results) >= 1

    def test_export_integration(self, sample_articles):
        skill = SkinMicrobiomeSkill(config_path="config.yaml")
        with tempfile.TemporaryDirectory() as tmpdir:
            fp = os.path.join(tmpdir, "integration.xlsx")
            result = skill.export(sample_articles, format="xlsx", output_path=fp)
            assert result == fp
            assert os.path.isfile(fp)
            assert os.path.getsize(fp) > 0

    def test_run_pipeline(self, entrez_mock):
        skill = SkinMicrobiomeSkill(config_path="config.yaml")
        with tempfile.TemporaryDirectory() as tmpdir:
            skill.exporter.output_dir = Path(tmpdir)
            articles = skill.run_pipeline(
                species="马",
                export_format="xlsx",
            )
            assert isinstance(articles, list)
            assert len(articles) >= 1
            for art in articles:
                assert art["species"] == "Equus caballus"
                assert "sampling_site" in art
                assert "bioproject_ids" in art

    def test_unknown_species_raises(self):
        skill = SkinMicrobiomeSkill(config_path="config.yaml")
        with pytest.raises(ValueError, match="Unknown species"):
            skill.run_pipeline(species="不存在的物种XYZ")

    def test_version(self):
        from skill import __version__
        assert __version__ == "0.3.0"
