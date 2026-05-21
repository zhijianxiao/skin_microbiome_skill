"""Export literature data to formatted Excel (.xlsx), CSV, or JSON.

Fixed 8-column output:
  文献标题 | 作者 | 年份 | DOI | 实验物种 | 宏基因组数据集 | 取样部位 | 相关度评分
Missing values → "N/A"
"""

from pathlib import Path
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ── Column definitions: (data_key, header_label) ───────────────────────
_COLUMNS = [
    ("title",            "文献标题"),
    ("authors",          "作者"),
    ("year",             "年份"),
    ("doi",              "DOI"),
    ("species",          "实验物种"),
    ("bioproject_ids",   "宏基因组数据集"),
    ("sampling_site",    "取样部位"),
    ("relevance_score",  "相关度评分"),
]

_NA = "N/A"

# ── Styling ────────────────────────────────────────────────────────────
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_FONT = Font(name="Microsoft YaHei", bold=True, color="FFFFFF", size=11)
_CELL_FONT   = Font(name="Microsoft YaHei", size=10)
_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
_COL_WIDTHS = [52, 28, 8, 26, 22, 32, 18, 10]


class Exporter:
    """Export results to formatted Excel (.xlsx), CSV, or JSON."""

    def __init__(self, config: dict):
        self.config = config
        ec = config.get("export", {})
        self.default_format = ec.get("default_format", "xlsx")
        self.output_dir = Path(ec.get("output_dir", "./output"))

    # ── dispatch ───────────────────────────────────────────────────────
    def export(
        self,
        data: list[dict],
        format: str = None,
        output_path: str = None,
        species_name: str = "",
    ) -> str | None:
        """Export data to file. Returns the output file path."""
        fmt = format or self.default_format

        if output_path:
            out = Path(output_path)
        else:
            self.output_dir.mkdir(parents=True, exist_ok=True)
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            safe = species_name.replace(" ", "_")[:40] if species_name else "results"
            out = self.output_dir / f"{safe}_{ts}.{fmt}"

        if fmt == "xlsx":
            return self.to_excel(data, str(out))
        elif fmt == "csv":
            return self.to_csv(data, str(out))
        elif fmt == "json":
            return self.to_json(data, str(out))
        else:
            raise ValueError(f"Unsupported format: {fmt}")

    # ── Excel ──────────────────────────────────────────────────────────
    def to_excel(self, data: list[dict], filepath: str) -> str:
        """Write a formatted .xlsx with auto-filter and freeze panes."""
        wb = Workbook()
        ws = wb.active
        ws.title = "文献检索结果"

        # Header row
        for col_idx, (_, header) in enumerate(_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _THIN_BORDER

        # Data rows
        for row_idx, article in enumerate(data, 2):
            row_values = self._build_row(article)
            for col_idx, value in enumerate(row_values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = _CELL_FONT
                cell.border = _THIN_BORDER
                if col_idx == 1:
                    cell.alignment = Alignment(wrap_text=True)

        # Column widths
        for col_idx, w in enumerate(_COL_WIDTHS, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = w

        # Freeze header + auto-filter
        ws.freeze_panes = "A2"
        last_col = get_column_letter(len(_COLUMNS))
        ws.auto_filter.ref = f"A1:{last_col}{max(len(data) + 1, 2)}"

        wb.save(filepath)
        return filepath

    # ── CSV ────────────────────────────────────────────────────────────
    def to_csv(self, data: list[dict], filepath: str) -> str:
        import csv
        headers = [h for _, h in _COLUMNS]
        with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            for article in data:
                writer.writerow(self._build_row(article))
        return filepath

    # ── JSON ───────────────────────────────────────────────────────────
    def to_json(self, data: list[dict], filepath: str) -> str:
        import json
        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        return filepath

    # ── helpers ────────────────────────────────────────────────────────
    def _build_row(self, article: dict) -> list:
        """Convert an article dict into a row list, N/A for missing fields."""
        bio = article.get("bioproject_ids", [])
        if isinstance(bio, list):
            bio_str = "; ".join(bio) if bio else _NA
        else:
            bio_str = str(bio) if bio else _NA

        species = article.get("species", _NA)
        if not species:
            species = _NA

        score = article.get("relevance_score", _NA)
        if score == "" or score is None:
            score = _NA

        return [
            article.get("title") or _NA,
            article.get("authors") or _NA,
            article.get("year") or _NA,
            article.get("doi") or _NA,
            species,
            bio_str,
            article.get("sampling_site") or _NA,
            score,
        ]
